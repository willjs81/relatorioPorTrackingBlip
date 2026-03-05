import uuid
import requests
import json
import re
import time
import concurrent.futures
from urllib.parse import quote
from openpyxl import Workbook
from datetime import datetime
from zoneinfo import ZoneInfo
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

URL = os.getenv("URL")
KEY_DUDA_PRD = os.getenv("KEY_DUDA_PRD")
KEY_FGTS_PRD = os.getenv("KEY_FGTS_PRD")
KEY_CLT_PRD = os.getenv("KEY_CLT_PRD")
KEY_DUDA_DEV = os.getenv("KEY_DUDA_DEV")


def limpar_mensagem(msg):
    """Remove quebras de linha e espacos duplicados para o resumo."""
    if not isinstance(msg, str):
        return msg
    msg = msg.replace("\n", " ").replace("\r", " ")
    msg = re.sub(r"\s{2,}", " ", msg)
    return msg.strip()

def normalizar_data(data_input: str) -> str | None:
    s = (data_input or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None

def normalizar_action(action_input: str) -> str:
    s = (action_input or "").strip()

    if len(s) >= 2 and (
        (s.startswith('"') and s.endswith('"')) or
        (s.startswith("'") and s.endswith("'"))
    ):
        s = s[1:-1].strip()

    if '\\"' in s:
        s = s.replace('\\"', '"')

    try:
        parsed = json.loads(s)
        if isinstance(parsed, dict):
            return json.dumps(parsed, separators=(",", ":"), ensure_ascii=False)
        if isinstance(parsed, str):
            parsed2 = json.loads(parsed)
            return json.dumps(parsed2, separators=(",", ":"), ensure_ascii=False)
    except Exception:
        pass

    return s


def gerar_candidatos_action(action_raw: str) -> list[str]:
    """
    Gera variacoes de action para consulta no endpoint /event-track/{tracking}/{action}.
    Alguns ambientes retornam action simples (ex: 1), mas o detalhamento pode exigir
    a representacao JSON string (ex: "1").
    """
    candidatos = []
    seen = set()

    def add(v):
        if not isinstance(v, str):
            return
        vv = v.strip()
        if not vv or vv in seen:
            return
        seen.add(vv)
        candidatos.append(vv)

    original = (action_raw or "").strip()
    add(original)

    normalizada = normalizar_action(action_raw)
    add(normalizada)

    # fallback principal para actions simples (ex.: 1 -> "1")
    add(json.dumps(original, ensure_ascii=False))

    # Se vier JSON escalar (numero/bool/null/string), tenta sua forma canonica.
    try:
        parsed = json.loads(original)
        if not isinstance(parsed, (dict, list)):
            add(json.dumps(parsed, ensure_ascii=False, separators=(",", ":")))
            if not isinstance(parsed, str):
                add(json.dumps(str(parsed), ensure_ascii=False))
    except Exception:
        pass

    return candidatos


def storage_date_br(storage_date_str: str) -> str:
    if not storage_date_str:
        return ""
    try:
        iso_str = storage_date_str.replace("Z", "+00:00")
        dt_utc = datetime.fromisoformat(iso_str)
        dt_br = dt_utc.astimezone(ZoneInfo("America/Sao_Paulo"))
        return dt_br.isoformat().replace("+00:00", "Z")
    except Exception:
        return storage_date_str


def parse_action_fields(action_str: str):
    """
    Extrai campos de action_raw e devolve (status, status_code, erro_message, action_dict).
    action_dict retorna None se action_raw nao for JSON dict.
    """
    status = None
    status_code = None
    erro_message = None
    action_dict = None

    try:
        parsed = json.loads(action_str)
        if isinstance(parsed, dict):
            action_dict = parsed
            status = parsed.get("status")
            status_code = parsed.get("statusCode")
            erro_field = parsed.get("erro")

            if isinstance(erro_field, dict):
                erro_message = erro_field.get("message")
            else:
                erro_message = erro_field
        else:
            erro_message = None
    except Exception:
        erro_message = None

    return status, status_code, erro_message, action_dict


def parse_json_object(value):
    """
    Se value for str/dict/list e puder ser interpretado como dict ou list, retorna o objeto.
    Caso contrario, retorna None.
    """
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, (dict, list)):
                return parsed
        except Exception:
            return None
    return None


def flatten_json(value, prefix: str = "") -> dict:
    """
    Achata (flatten) objetos JSON em um dict de chaves -> valores.
    - Dicts viram prefix_key_subkey...
    - Lists viram prefix_0_subkey...
    - Strings que forem JSON (dict/list) tambem sao expandidas.
    """
    if value is None:
        return {}

    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, (dict, list)):
                return flatten_json(parsed, prefix)
        except Exception:
            pass
        return {prefix: value} if prefix else {}

    if isinstance(value, dict):
        out = {}
        for key, item in value.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.update(flatten_json(item, next_prefix))
        return out

    if isinstance(value, list):
        out = {}
        for idx, item in enumerate(value):
            next_prefix = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
            out.update(flatten_json(item, next_prefix))
        if not value and prefix:
            out[prefix] = "[]"
        return out

    return {prefix: value} if prefix else {}


def sanitize_column_name(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-z_]+", "_", str(name))
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def router_slug(nome: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z]+", "_", nome.strip()).strip("_").lower()
    return slug or "router"


def escolher_autorizacao():
    print("\nSelecione o router para usar a chave de autorizacao:")
    print("1 - Duda PRD")
    print("2 - Campanhas FGTS PRD (padrao)")
    print("3 - Campanhas Emp Privado")
    print("4 - Duda DEV")

    opcao = input("Opcao (1/2/3/4, ENTER = 2): ").strip()

    if opcao == "1":
        token = KEY_DUDA_PRD
        nome = "Duda PRD"
    elif opcao == "3":
        token = KEY_CLT_PRD
        nome = "Campanhas Emp Privado"
    elif opcao == "4":
        token = KEY_DUDA_DEV
        nome = "Duda DEV"
    else:
        token = KEY_FGTS_PRD
        nome = "Campanhas FGTS PRD"

    if not token:
        print(f"\nChave nao encontrada para {nome}. Verifique o .env.\n")
        return None, None, None

    print(f"\nUsando chave do router: {nome}")
    return token, router_slug(nome), nome


def fetch_paginated_items(uri_base: str, headers: dict, take: int, max_retries: int, contexto: str):
    """Busca todas as paginas da API usando $take/$skip."""
    all_items = []
    skip = 0
    total_esperado = None
    page = 0

    while True:
        page += 1
        sep = "&" if "?" in uri_base else "?"
        uri = f"{uri_base}{sep}$take={take}&$skip={skip}"
        body = {
            "id": str(uuid.uuid4()),
            "to": "postmaster@analytics.msging.net",
            "method": "get",
            "uri": uri,
        }

        last_error = ""
        last_status_code = ""
        last_body = ""
        response = None
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.post(URL, json=body, headers=headers, timeout=30)
                response.raise_for_status()
                last_error = ""
                last_status_code = ""
                last_body = ""
                break
            except requests.RequestException as err:
                if response is not None:
                    last_status_code = str(response.status_code)
                    last_error = f"HTTP {response.status_code}"
                else:
                    last_status_code = ""
                    last_error = f"{err.__class__.__name__}: {err}"
                try:
                    last_body = response.text if response is not None else ""
                except Exception:
                    last_body = ""
                if attempt < max_retries:
                    print(
                        f"[{contexto}] erro na pagina {page} ({last_error}). "
                        f"Tentando novamente ({attempt}/{max_retries})..."
                    )
                    time.sleep(1.2)

        if last_error:
            return {
                "failed": {
                    "erro": last_error,
                    "status_code": last_status_code,
                    "response_body": last_body,
                },
                "items": all_items,
                "partial": bool(all_items),
            }

        data = response.json()
        resource = data.get("resource", {})
        items = resource.get("items", [])
        if total_esperado is None:
            total_esperado = resource.get("total")

        all_items.extend(items)
        print(f"[{contexto}] pagina {page} -> {len(items)} item(s) (acumulado: {len(all_items)})")

        if not items:
            break
        if len(items) < take:
            break
        if isinstance(total_esperado, int) and len(all_items) >= total_esperado:
            break

        skip += take

    return {"items": all_items}


while True:
    print("\n===== CONSULTA DE EVENTOS BLiP (ANALYTICS) =====")

    # Entrada do usuario
    start_date_input = input("Informe a data inicial (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    end_date_input = input("Informe a data final   (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    start_date = normalizar_data(start_date_input)
    end_date = normalizar_data(end_date_input)

    if not start_date or not end_date:
        print("\nData invalida. Use DD-MM-AAAA ou YYYY-MM-DD.\n")
        continue

    default_tracking = "md Response api inclusao response tratada"
    tracking_input = input(
        f'Digite o nome do tracking (ENTER para usar o padrao "{default_tracking}"): '
    ).strip()

    tracking = tracking_input if tracking_input else default_tracking
    tracking_encoded = quote(tracking, safe="")

    uri_event = f"/event-track/{tracking_encoded}?startDate={start_date}&endDate={end_date}"

    print("\n===== PARAMETROS DA CONSULTA =====")
    print(f"Data inicial : {start_date}")
    print(f"Data final   : {end_date}")
    print(f"Tracking     : {tracking}")
    print(f"URI montada  : {uri_event}")

    authorization, router_dir, router_nome = escolher_autorizacao()
    if not authorization:
        continue

    headers = {"Authorization": authorization, "Content-Type": "application/json"}

    print("\n===== STATUS DA REQUISICAO =====")
    print("Paginando /event-track com $take=1000")
    result_main = fetch_paginated_items(
        uri_base=uri_event,
        headers=headers,
        take=1000,
        max_retries=3,
        contexto="event-track",
    )
    failed_main = result_main.get("failed")
    if failed_main:
        print("\nErro na requisicao! Voltando ao inicio...\n")
        print(f"Detalhe: {failed_main.get('erro')} (status={failed_main.get('status_code')})")
        continue

    items = result_main.get("items", [])

    print("\n===== QUANTIDADE DE REGISTROS ENCONTRADOS =====")
    print(len(items))

    if len(items) == 0:
        print("\nNenhum registro encontrado. Voltando ao inicio...\n")
        continue

    # Ordena registros por count
    items_sorted = sorted(items, key=lambda x: x.get("count", 0), reverse=True)

    # Lista itens ordenados no console
    print("\n===== ITEMS ORDENADOS POR COUNT =====")
    print(json.dumps(items_sorted, indent=4, ensure_ascii=False))

    # Prepara dados enriquecidos e agrega inconsistencias para uso no Excel
    erros_agrupados = {}
    items_enriquecidos = []
    flat_keys = set()
    action_info = {}

    for item in items_sorted:
        action_str = item.get("action", "")
        count = item.get("count", 0)
        status_default = "sem status"
        status_code_default = "sem codigo"

        status, status_code, erro_message, action_dict = parse_action_fields(action_str)
        extras_obj = parse_json_object(item.get("extras"))

        action_flat_raw = flatten_json(action_dict, "action") if action_dict else {}
        extras_flat_raw = flatten_json(extras_obj, "extras") if extras_obj is not None else {}

        action_flat = {sanitize_column_name(k): v for k, v in action_flat_raw.items()}
        extras_flat = {sanitize_column_name(k): v for k, v in extras_flat_raw.items()}

        mensagem = erro_message or f"{status or status_default} (statusCode={status_code or status_code_default})"
        erros_agrupados[mensagem] = erros_agrupados.get(mensagem, 0) + count

        flat_keys.update(action_flat.keys())
        flat_keys.update(extras_flat.keys())

        items_enriquecidos.append(
            {
                "item": item,
                "status": status,
                "status_code": status_code,
                "erro_message": erro_message,
                "action_dict": action_dict,
                "extras_obj": extras_obj,
                "action_flat": action_flat,
                "extras_flat": extras_flat,
            }
        )

        action_raw = item.get("action", "")
        info = action_info.setdefault(
            action_raw,
            {
                "storage_dates": set(),
                "status": set(),
                "status_code": set(),
                "erro_message": set(),
            },
        )
        if item.get("storageDate"):
            info["storage_dates"].add(item.get("storageDate"))
        if status:
            info["status"].add(str(status))
        if status_code:
            info["status_code"].add(str(status_code))
        if erro_message:
            info["erro_message"].add(str(erro_message))

    erros_ordenados = sorted(erros_agrupados.items(), key=lambda x: x[1], reverse=True)

    # Pergunta se gera Excel
    gerar = input("\nDeseja gerar o relatorio em Excel (s/n)? ").strip().lower()

    if gerar != "s":
        print("\nRelatorio NAO gerado. Encerrando o script.\n")
        break

    # --------------------------------------------------------------------
    # COLETA EVENTOS POR ACTION (CONSOLIDADO)
    # --------------------------------------------------------------------
    print("\n===== COLETANDO EVENTOS POR ACTION =====")
    eventos_rows = []
    extras_keys_eventos = set()
    failed_actions = []
    max_retries = 3
    max_workers = 5

    unique_actions = []
    seen_actions = set()
    for enriched in items_enriquecidos:
        action_raw = enriched["item"].get("action", "")
        if action_raw not in seen_actions:
            seen_actions.add(action_raw)
            unique_actions.append(action_raw)

    def fetch_action(action_raw: str):
        candidatos = gerar_candidatos_action(action_raw)
        resultado_sucesso = None
        ultima_falha = None
        action_normalizada = normalizar_action(action_raw)

        for idx_candidato, action_candidata in enumerate(candidatos, start=1):
            action_encoded = quote(action_candidata, safe="")
            uri_event_action_base = (
                f"/event-track/{tracking_encoded}/"
                f"{action_encoded}"
                f"?startDate={start_date}&endDate={end_date}"
            )

            print(
                f"[action] tentativa {idx_candidato}/{len(candidatos)} "
                f"raw={action_raw!r} consulta={action_candidata!r}"
            )
            result_action = fetch_paginated_items(
                uri_base=uri_event_action_base,
                headers=headers,
                take=500,
                max_retries=max_retries,
                contexto=f"action:{action_raw[:40]}",
            )
            failed = result_action.get("failed")
            if failed:
                itens_parciais = result_action.get("items", [])
                if itens_parciais:
                    print(
                        f"[action] retorno parcial com consulta={action_candidata!r} "
                        f"-> {len(itens_parciais)} item(s)"
                    )
                    resultado_sucesso = {"items": itens_parciais, "action_consultada": action_candidata}
                    break
                ultima_falha = {
                    "action_raw": action_raw,
                    "action_normalizada": action_normalizada,
                    "action_consultada": action_candidata,
                    "erro": failed.get("erro", ""),
                    "status_code": failed.get("status_code", ""),
                    "response_body": failed.get("response_body", ""),
                }
                continue

            itens = result_action.get("items", [])
            if itens:
                print(f"[action] sucesso com consulta={action_candidata!r} -> {len(itens)} item(s)")
                resultado_sucesso = {"items": itens, "action_consultada": action_candidata}
                break

            # guarda resultado vazio para usar caso nenhum candidato retorne itens
            if resultado_sucesso is None:
                resultado_sucesso = {"items": [], "action_consultada": action_candidata}

        if resultado_sucesso is None and ultima_falha:
            return {"failed": ultima_falha}

        items_action = (resultado_sucesso or {}).get("items", [])

        rows = []
        extras_keys = set()
        vistos = set()
        for item in items_action:
            contact = item.get("contact", {}).get("Identity")
            action_item = item.get("action")
            chave = (contact, action_item)
            if chave in vistos:
                continue
            vistos.add(chave)

            extras = item.get("extras") or {}
            extras_keys.update(extras.keys())

            info = action_info.get(action_raw, {})
            rows.append(
                {
                    "inc_status": ";".join(sorted(info.get("status", set()))),
                    "inc_status_code": ";".join(sorted(info.get("status_code", set()))),
                    "inc_erro_message": " | ".join(sorted(info.get("erro_message", set()))),
                    "storageDate": item.get("storageDate"),
                    "storageDateBR": storage_date_br(item.get("storageDate")),
                    "cpf": extras.get("cpf", ""),
                    "contactIdentity": contact or "",
                    "category": item.get("category"),
                    "action": action_item,
                    "extras": extras,
                }
            )

        return {
            "rows": rows,
            "extras_keys": extras_keys,
        }

    total_actions = len(unique_actions)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(fetch_action, action_raw): (idx, action_raw)
            for idx, action_raw in enumerate(unique_actions, start=1)
        }
        for future in concurrent.futures.as_completed(future_map):
            idx, action_raw_future = future_map[future]
            print(f"[{idx}/{total_actions}] Consultando action...")
            try:
                result = future.result()
            except Exception as err:
                failed_actions.append(
                    {
                        "action_raw": action_raw_future,
                        "action_normalizada": normalizar_action(action_raw_future),
                        "erro": f"{err.__class__.__name__}: {err}",
                        "status_code": "",
                        "response_body": "",
                    }
                )
                continue
            failed = result.get("failed")
            if failed:
                failed_actions.append(failed)
                continue
            eventos_rows.extend(result.get("rows", []))
            extras_keys_eventos.update(result.get("extras_keys", set()))

    # Criacao do Excel com 2 abas
    wb = Workbook()

    # Aba 1 - inconsistencias_detalhadas
    ws_det = wb.active
    ws_det.title = "relatorio por eventos"

    ws_det.append(["Router", router_nome])
    ws_det.append(["Tracking", tracking])
    ws_det.append(["Data inicial", start_date])
    ws_det.append(["Data final", end_date])
    ws_det.append([])
    ws_det.append(
        [
            "storageDate",
            "category",
            "status",
            "statusCode",
            "erro_message",
            "count",
            "action_raw",
        ]
    )

    # Usa items_enriquecidos para evitar reprocessar action_raw
    for enriched in items_enriquecidos:
        item = enriched["item"]
        storage_date = item.get("storageDate")
        category = item.get("category")
        count = item.get("count", 0)
        action_str = item.get("action", "")
        status = enriched["status"]
        status_code = enriched["status_code"]
        erro_message = enriched["erro_message"]

        ws_det.append(
            [
                storage_date,
                category,
                status,
                status_code,
                erro_message,
                count,
                action_str,
            ]
        )

    # Aba 2 - eventos_por_action (consolidado)
    ws_eventos = wb.create_sheet("relatorio por action")

    extras_headers = [f"Extras_{key}" for key in sorted(extras_keys_eventos)]
    ws_eventos.append(
        [
            "inc_status",
            "inc_statusCode",
            "inc_erro_message",
            "storageDate",
            "storageDateBR",
            "cpf",
            "contactIdentity",
            "category",
            "action",
            *extras_headers,
        ]
    )

    for row in eventos_rows:
        extras = row.get("extras") or {}
        extras_values = []
        for key in sorted(extras_keys_eventos):
            value = extras.get(key, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            if value is None:
                value = ""
            extras_values.append(value)

        ws_eventos.append(
            [
                row.get("inc_status", ""),
                row.get("inc_status_code", ""),
                row.get("inc_erro_message", ""),
                row.get("storageDate", ""),
                row.get("storageDateBR", ""),
                row.get("cpf", ""),
                row.get("contactIdentity", ""),
                row.get("category", ""),
                row.get("action", ""),
                *extras_values,
            ]
        )

    # Aba 3 - actions_falharam (se houver)
    if failed_actions:
        ws_failed = wb.create_sheet("actions falharam")
        ws_failed.append(["action_raw", "action_normalizada", "erro", "status_code", "response_body"])
        for item in failed_actions:
            ws_failed.append(
                [
                    item.get("action_raw", ""),
                    item.get("action_normalizada", ""),
                    item.get("erro", ""),
                    item.get("status_code", ""),
                    item.get("response_body", ""),
                ]
            )

    today_str = datetime.now().strftime("%Y%m%d")

    home_dir = Path.home()
    documentos_dir = home_dir / "Documentos"
    if not documentos_dir.exists():
        documentos_dir = home_dir / "Documents"

    relatorio_root = documentos_dir / "relatorio de eventos"
    tipo_subpasta = relatorio_root / "eventDetails" / router_dir
    today_folder = tipo_subpasta / f"relatorio_{today_str}"
    today_folder.mkdir(parents=True, exist_ok=True)

    base_name = f"relatorio_consolidado_eventos_{router_dir}_{today_str}"
    ext = ".xlsx"
    file_name = base_name + ext
    file_path = today_folder / file_name

    version = 1
    while file_path.exists():
        version += 1
        file_name = f"{base_name}_v{version}{ext}"
        file_path = today_folder / file_name

    wb.save(file_path)

    print("\n===== ARQUIVO GERADO =====")
    print("Arquivo salvo como:", file_path)
    if failed_actions:
        print("\n===== RESUMO DE FALHAS =====")
        print(f"Actions com erro: {len(failed_actions)}")
        print("Confira a aba 'actions_falharam' no Excel para detalhes.")
    print("\nEncerrando o programa.\n")
    break
