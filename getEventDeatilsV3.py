import uuid
import requests
import json
import re
from urllib.parse import quote
from openpyxl import Workbook
from datetime import datetime
import os
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

URL = os.getenv("URL")
KEY_DUDA_PRD = os.getenv("KEY_DUDA_PRD")
KEY_FGTS_PRD = os.getenv("KEY_FGTS_PRD")
KEY_CLT_PRD = os.getenv("KEY_CLT_PRD")
KEY_DUDA_DEV = os.getenv("KEY_DUDA_DEV")


def limpar_mensagem(msg):
    """Remove quebras de linha e espacos duplicados para o resumo."""
    if not isinstance(msg, str):
        return msg
    msg = msg.replace("\n", " ").replace("\r", " ")
    msg = re.sub(r"\s{2,}", " ", msg)
    return msg.strip()

def normalizar_data(data_input: str) -> str | None:
    s = (data_input or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def parse_action_fields(action_str: str):
    """
    Extrai campos de action_raw e devolve (status, status_code, erro_message, action_dict).
    action_dict retorna None se action_raw nao for JSON dict.
    """
    status = None
    status_code = None
    erro_message = None
    action_dict = None

    try:
        parsed = json.loads(action_str)
        if isinstance(parsed, dict):
            action_dict = parsed
            status = parsed.get("status")
            status_code = parsed.get("statusCode")
            erro_field = parsed.get("erro")

            if isinstance(erro_field, dict):
                erro_message = erro_field.get("message")
            else:
                erro_message = erro_field
        else:
            erro_message = str(parsed)
    except Exception:
        erro_message = action_str.strip() or "ACTION_NAO_JSON"

    return status, status_code, erro_message, action_dict


def parse_json_object(value):
    """
    Se value for str/dict/list e puder ser interpretado como dict ou list, retorna o objeto.
    Caso contrario, retorna None.
    """
    if isinstance(value, dict):
        return value
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, (dict, list)):
                return parsed
        except Exception:
            return None
    return None


def flatten_json(value, prefix: str = "") -> dict:
    """
    Achata (flatten) objetos JSON em um dict de chaves -> valores.
    - Dicts viram prefix_key_subkey...
    - Lists viram prefix_0_subkey...
    - Strings que forem JSON (dict/list) tambem sao expandidas.
    """
    if value is None:
        return {}

    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            if isinstance(parsed, (dict, list)):
                return flatten_json(parsed, prefix)
        except Exception:
            pass
        return {prefix: value} if prefix else {}

    if isinstance(value, dict):
        out = {}
        for key, item in value.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            out.update(flatten_json(item, next_prefix))
        return out

    if isinstance(value, list):
        out = {}
        for idx, item in enumerate(value):
            next_prefix = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
            out.update(flatten_json(item, next_prefix))
        if not value and prefix:
            out[prefix] = "[]"
        return out

    return {prefix: value} if prefix else {}


def sanitize_column_name(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-z_]+", "_", str(name))
    name = re.sub(r"_+", "_", name).strip("_")
    return name


def router_slug(nome: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z]+", "_", nome.strip()).strip("_").lower()
    return slug or "router"


def escolher_autorizacao():
    print("\nSelecione o router para usar a chave de autorizacao:")
    print("1 - Duda PRD")
    print("2 - Campanhas FGTS PRD (padrao)")
    print("3 - Campanhas Emp Privado")
    print("4 - Duda DEV")

    opcao = input("Opcao (1/2/3/4, ENTER = 2): ").strip()

    if opcao == "1":
        token = KEY_DUDA_PRD
        nome = "Duda PRD"
    elif opcao == "3":
        token = KEY_CLT_PRD
        nome = "Campanhas Emp Privado"
    elif opcao == "4":
        token = KEY_DUDA_DEV
        nome = "Duda DEV"
    else:
        token = KEY_FGTS_PRD
        nome = "Campanhas FGTS PRD"

    if not token:
        print(f"\nChave nao encontrada para {nome}. Verifique o .env.\n")
        return None, None, None

    print(f"\nUsando chave do router: {nome}")
    return token, router_slug(nome), nome


while True:
    print("\n===== CONSULTA DE EVENTOS BLiP (ANALYTICS) =====")

    # Entrada do usuario
    start_date_input = input("Informe a data inicial (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    end_date_input = input("Informe a data final   (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    start_date = normalizar_data(start_date_input)
    end_date = normalizar_data(end_date_input)

    if not start_date or not end_date:
        print("\nData invalida. Use DD-MM-AAAA ou YYYY-MM-DD.\n")
        continue

    default_tracking = "md Response api inclusao response tratada"
    tracking_input = input(
        f'Digite o nome do tracking (ENTER para usar o padrao "{default_tracking}"): '
    ).strip()

    tracking = tracking_input if tracking_input else default_tracking
    tracking_encoded = quote(tracking, safe="")

    uri_event = (
        f"/event-track/{tracking_encoded}"
        f"?startDate={start_date}&endDate={end_date}&$take=1000"
    )

    print("\n===== PARAMETROS DA CONSULTA =====")
    print(f"Data inicial : {start_date}")
    print(f"Data final   : {end_date}")
    print(f"Tracking     : {tracking}")
    print(f"URI montada  : {uri_event}")

    authorization, router_dir, router_nome = escolher_autorizacao()
    if not authorization:
        continue

    headers = {"Authorization": authorization, "Content-Type": "application/json"}

    body = {
        "id": str(uuid.uuid4()),
        "to": "postmaster@analytics.msging.net",
        "method": "get",
        "uri": uri_event,
    }

    response = requests.post(URL, json=body, headers=headers)

    print("\n===== STATUS DA REQUISICAO =====")
    print(response.status_code)

    try:
        response.raise_for_status()
    except requests.HTTPError:
        print("\nErro na requisicao! Voltando ao inicio...\n")
        continue

    data = response.json()
    items = data.get("resource", {}).get("items", [])

    print("\n===== QUANTIDADE DE REGISTROS ENCONTRADOS =====")
    print(len(items))

    if len(items) == 0:
        print("\nNenhum registro encontrado. Voltando ao inicio...\n")
        continue

    # Ordena registros por count
    items_sorted = sorted(items, key=lambda x: x.get("count", 0), reverse=True)

    # Lista itens ordenados no console
    print("\n===== ITEMS ORDENADOS POR COUNT =====")
    print(json.dumps(items_sorted, indent=4, ensure_ascii=False))

    # Prepara dados enriquecidos e agrega inconsistencias para uso no Excel
    erros_agrupados = {}
    items_enriquecidos = []
    flat_keys = set()

    for item in items_sorted:
        action_str = item.get("action", "")
        count = item.get("count", 0)
        status_default = "sem status"
        status_code_default = "sem codigo"

        status, status_code, erro_message, action_dict = parse_action_fields(action_str)
        extras_obj = parse_json_object(item.get("extras"))

        action_flat_raw = flatten_json(action_dict, "action") if action_dict else {}
        extras_flat_raw = flatten_json(extras_obj, "extras") if extras_obj is not None else {}

        action_flat = {sanitize_column_name(k): v for k, v in action_flat_raw.items()}
        extras_flat = {sanitize_column_name(k): v for k, v in extras_flat_raw.items()}

        mensagem = erro_message or f"{status or status_default} (statusCode={status_code or status_code_default})"
        erros_agrupados[mensagem] = erros_agrupados.get(mensagem, 0) + count

        flat_keys.update(action_flat.keys())
        flat_keys.update(extras_flat.keys())

        items_enriquecidos.append(
            {
                "item": item,
                "status": status,
                "status_code": status_code,
                "erro_message": erro_message,
                "action_dict": action_dict,
                "extras_obj": extras_obj,
                "action_flat": action_flat,
                "extras_flat": extras_flat,
            }
        )


    erros_ordenados = sorted(erros_agrupados.items(), key=lambda x: x[1], reverse=True)

    # Pergunta se gera Excel
    gerar = input("\nDeseja gerar o relatorio em Excel (s/n)? ").strip().lower()

    if gerar != "s":
        print("\nRelatorio NAO gerado. Encerrando o script.\n")
        break

    # Criacao do Excel com 2 abas
    wb = Workbook()

    # Aba 1 - inconsistencias_detalhadas
    ws_det = wb.active
    ws_det.title = "relatorio por eventos"

    ws_det.append(["Router", router_nome])
    ws_det.append(["Tracking", tracking])
    ws_det.append(["Data inicial", start_date])
    ws_det.append(["Data final", end_date])
    ws_det.append([])
    ws_det.append(
        [
            "storageDate",
            "category",
            "status",
            "statusCode",
            "erro_message",
            "count",
            "action_raw",
        ]
    )

    # Usa items_enriquecidos para evitar reprocessar action_raw
    for enriched in items_enriquecidos:
        item = enriched["item"]
        storage_date = item.get("storageDate")
        category = item.get("category")
        count = item.get("count", 0)
        action_str = item.get("action", "")
        status = enriched["status"]
        status_code = enriched["status_code"]
        erro_message = enriched["erro_message"]

        ws_det.append(
            [
                storage_date,
                category,
                status,
                status_code,
                erro_message,
                count,
                action_str,
            ]
        )

    # Aba 2 - resumo_inconsistencias
    ws_resumo = wb.create_sheet("resumo inconsistencias")

    ws_resumo.append(["Router", router_nome])
    ws_resumo.append(["Tracking", tracking])
    ws_resumo.append(["Data inicial", start_date])
    ws_resumo.append(["Data final", end_date])
    ws_resumo.append([])
    ws_resumo.append(["ocorrencias_totais", "erro_message"])

    for mensagem, total in erros_ordenados:
        mensagem_limpa = limpar_mensagem(mensagem)
        ws_resumo.append([total, mensagem_limpa])

    # Aba 3 - action_raw_expandido
    ws_action = wb.create_sheet("action raw expandido")

    flat_headers = sorted(flat_keys)
    ws_action.append(["Router", router_nome])
    ws_action.append(["Tracking", tracking])
    ws_action.append(["Data inicial", start_date])
    ws_action.append(["Data final", end_date])
    ws_action.append([])
    ws_action.append(
        [
            "storageDate",
            "category",
            "status",
            "statusCode",
            "erro_message",
            "count",
            "action_raw",
            "extras_raw",
            *flat_headers,
        ]
    )

    for enriched in items_enriquecidos:
        item = enriched["item"]
        action_flat = enriched["action_flat"] or {}
        extras_flat = enriched["extras_flat"] or {}
        storage_date = item.get("storageDate")
        category = item.get("category")
        count = item.get("count", 0)
        action_str = item.get("action", "")
        extras_raw = item.get("extras", "")
        if isinstance(extras_raw, (dict, list)):
            extras_raw = json.dumps(extras_raw, ensure_ascii=False)
        status = enriched["status"]
        status_code = enriched["status_code"]
        erro_message = enriched["erro_message"]

        flat_values = []
        for col_name in flat_headers:
            value = action_flat.get(col_name, "")
            if value == "":
                value = extras_flat.get(col_name, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            if value is None:
                value = ""
            flat_values.append(value)

        ws_action.append(
            [
                storage_date,
                category,
                status,
                status_code,
                erro_message,
                count,
                action_str,
                extras_raw,
                *flat_values,
            ]
        )


    today_str = datetime.now().strftime("%Y%m%d")

    home_dir = Path.home()
    documentos_dir = home_dir / "Documentos"
    if not documentos_dir.exists():
        documentos_dir = home_dir / "Documents"

    relatorio_root = documentos_dir / "relatorio de eventos"
    tipo_subpasta = relatorio_root / "eventDetails" / router_dir
    today_folder = tipo_subpasta / f"relatorio_{today_str}"
    today_folder.mkdir(parents=True, exist_ok=True)

    base_name = f"relatorio_eventos_bmg_{router_dir}_{today_str}"
    ext = ".xlsx"
    file_name = base_name + ext
    file_path = today_folder / file_name

    version = 1
    while file_path.exists():
        version += 1
        file_name = f"{base_name}_v{version}{ext}"
        file_path = today_folder / file_name

    wb.save(file_path)

    print("\n===== ARQUIVO GERADO =====")
    print("Arquivo salvo como:", file_path)
    print("\nEncerrando o programa.\n")
    break
