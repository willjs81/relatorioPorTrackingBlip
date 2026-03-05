import concurrent.futures
import json
import os
import re
import uuid
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()

URL = os.getenv("URL")
KEY_DUDA_PRD = os.getenv("KEY_DUDA_PRD")
KEY_FGTS_PRD = os.getenv("KEY_FGTS_PRD")
KEY_CLT_PRD = os.getenv("KEY_CLT_PRD")
KEY_DUDA_DEV = os.getenv("KEY_DUDA_DEV")


def limpar_mensagem(msg):
    if not isinstance(msg, str):
        return msg
    msg = msg.replace("\n", " ").replace("\r", " ")
    msg = re.sub(r"\s{2,}", " ", msg)
    return msg.strip()


def normalizar_data(data_input: str) -> str | None:
    s = (data_input or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def normalizar_action(action_input: str) -> str:
    s = (action_input or "").strip()

    if len(s) >= 2 and (
        (s.startswith('"') and s.endswith('"'))
        or (s.startswith("'") and s.endswith("'"))
    ):
        s = s[1:-1].strip()

    if '\\"' in s:
        s = s.replace('\\"', '"')

    try:
        parsed = json.loads(s)
        if isinstance(parsed, dict):
            return json.dumps(parsed, separators=(",", ":"), ensure_ascii=False)
        if isinstance(parsed, str):
            parsed2 = json.loads(parsed)
            return json.dumps(parsed2, separators=(",", ":"), ensure_ascii=False)
    except Exception:
        pass

    return s


def parse_action_fields(action_str: str):
    status = None
    status_code = None
    erro_message = None

    try:
        parsed = json.loads(action_str)
        if isinstance(parsed, dict):
            status = parsed.get("status")
            status_code = parsed.get("statusCode")
            erro_field = parsed.get("erro")
            if isinstance(erro_field, dict):
                erro_message = erro_field.get("message")
            else:
                erro_message = erro_field
        else:
            erro_message = str(parsed)
    except Exception:
        erro_message = action_str.strip() or "ACTION_NAO_JSON"

    return status, status_code, erro_message


def storage_date_br(storage_date_str: str) -> str:
    if not storage_date_str:
        return ""
    try:
        iso_str = storage_date_str.replace("Z", "+00:00")
        dt_utc = datetime.fromisoformat(iso_str)
        dt_br = dt_utc.astimezone(ZoneInfo("America/Sao_Paulo"))
        return dt_br.isoformat().replace("+00:00", "Z")
    except Exception:
        return storage_date_str


def router_slug(nome: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z]+", "_", nome.strip()).strip("_").lower()
    return slug or "router"


def escolher_autorizacao():
    print("\nSelecione o router para usar a chave de autorizacao:")
    print("1 - Duda PRD")
    print("2 - Campanhas FGTS PRD (padrao)")
    print("3 - Campanhas Emp Privado")
    print("4 - Duda DEV")

    opcao = input("Opcao (1/2/3/4, ENTER = 2): ").strip()

    if opcao == "1":
        token = KEY_DUDA_PRD
        nome = "Duda PRD"
    elif opcao == "3":
        token = KEY_CLT_PRD
        nome = "Campanhas Emp Privado"
    elif opcao == "4":
        token = KEY_DUDA_DEV
        nome = "Duda DEV"
    else:
        token = KEY_FGTS_PRD
        nome = "Campanhas FGTS PRD"

    if not token:
        print(f"\nChave nao encontrada para {nome}. Verifique o .env.\n")
        return None, None, None

    print(f"\nUsando chave do router: {nome}")
    return token, router_slug(nome), nome


def status_code_int(status_code) -> int | None:
    if status_code is None:
        return None
    try:
        return int(status_code)
    except Exception:
        return None


def is_error(status_code, erro_message) -> bool:
    code = status_code_int(status_code)
    if code is not None and code >= 400:
        return True
    if isinstance(erro_message, str) and erro_message.strip():
        return True
    return False


def date_key(iso_like: str) -> str:
    if not iso_like:
        return ""
    return str(iso_like)[:10]


while True:
    print("\n===== CONSULTA DE EVENTOS BLiP (ANALITICO) =====")

    start_date_input = input("Informe a data inicial (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    end_date_input = input("Informe a data final   (DD-MM-AAAA ou YYYY-MM-DD): ").strip()
    start_date = normalizar_data(start_date_input)
    end_date = normalizar_data(end_date_input)

    if not start_date or not end_date:
        print("\nData invalida. Use DD-MM-AAAA ou YYYY-MM-DD.\n")
        continue

    default_tracking = "md Response api inclusao response tratada"
    tracking_input = input(
        f'Digite o nome do tracking (ENTER para usar o padrao "{default_tracking}"): '
    ).strip()

    tracking = tracking_input if tracking_input else default_tracking
    tracking_encoded = quote(tracking, safe="")

    uri_event = (
        f"/event-track/{tracking_encoded}"
        f"?startDate={start_date}&endDate={end_date}&$take=1000"
    )

    print("\n===== PARAMETROS DA CONSULTA =====")
    print(f"Data inicial : {start_date}")
    print(f"Data final   : {end_date}")
    print(f"Tracking     : {tracking}")
    print(f"URI montada  : {uri_event}")

    authorization, router_dir, router_nome = escolher_autorizacao()
    if not authorization:
        continue

    headers = {"Authorization": authorization, "Content-Type": "application/json"}

    body = {
        "id": str(uuid.uuid4()),
        "to": "postmaster@analytics.msging.net",
        "method": "get",
        "uri": uri_event,
    }

    response = requests.post(URL, json=body, headers=headers)

    print("\n===== STATUS DA REQUISICAO =====")
    print(response.status_code)

    try:
        response.raise_for_status()
    except requests.HTTPError:
        print("\nErro na requisicao! Voltando ao inicio...\n")
        continue

    data = response.json()
    items = data.get("resource", {}).get("items", [])

    print("\n===== QUANTIDADE DE REGISTROS ENCONTRADOS =====")
    print(len(items))

    if len(items) == 0:
        print("\nNenhum registro encontrado. Voltando ao inicio...\n")
        continue

    items_sorted = sorted(items, key=lambda x: x.get("count", 0), reverse=True)

    print("\n===== ITEMS ORDENADOS POR COUNT =====")
    print(json.dumps(items_sorted, indent=4, ensure_ascii=False))

    erros_agrupados = {}
    items_enriquecidos = []
    action_info = {}

    for item in items_sorted:
        action_str = item.get("action", "")
        count = int(item.get("count", 0) or 0)
        status_default = "sem status"
        status_code_default = "sem codigo"

        status, status_code, erro_message = parse_action_fields(action_str)

        mensagem = erro_message or (
            f"{status or status_default} (statusCode={status_code or status_code_default})"
        )
        erros_agrupados[mensagem] = erros_agrupados.get(mensagem, 0) + count

        item_error = is_error(status_code, erro_message)

        items_enriquecidos.append(
            {
                "item": item,
                "status": status,
                "status_code": status_code,
                "erro_message": erro_message,
                "is_error": item_error,
            }
        )

        action_raw = item.get("action", "")
        info = action_info.setdefault(
            action_raw,
            {
                "status": set(),
                "status_code": set(),
                "erro_message": set(),
                "total_count": 0,
                "error_count": 0,
            },
        )
        if status:
            info["status"].add(str(status))
        if status_code is not None:
            info["status_code"].add(str(status_code))
        if erro_message:
            info["erro_message"].add(str(erro_message))
        info["total_count"] += count
        if item_error:
            info["error_count"] += count

    erros_ordenados = sorted(erros_agrupados.items(), key=lambda x: x[1], reverse=True)

    gerar = input("\nDeseja gerar o relatorio em Excel (s/n)? ").strip().lower()

    if gerar != "s":
        print("\nRelatorio NAO gerado. Encerrando o script.\n")
        break

    print("\n===== COLETANDO EVENTOS POR ACTION (SEM DEDUPLICAR) =====")
    eventos_rows = []
    extras_keys_eventos = set()
    failed_actions = []
    max_retries = 3
    max_workers = 10

    unique_actions = []
    seen_actions = set()
    for enriched in items_enriquecidos:
        action_raw = enriched["item"].get("action", "")
        if action_raw not in seen_actions:
            seen_actions.add(action_raw)
            unique_actions.append(action_raw)

    def fetch_action(action_raw: str):
        action_normalizada = normalizar_action(action_raw)
        action_encoded = quote(action_normalizada, safe="")

        uri_event_action = (
            f"/event-track/{tracking_encoded}/"
            f"{action_encoded}"
            f"?startDate={start_date}&endDate={end_date}&$take=500"
        )

        body_action = {
            "id": str(uuid.uuid4()),
            "to": "postmaster@analytics.msging.net",
            "method": "get",
            "uri": uri_event_action,
        }

        last_error = ""
        last_status_code = ""
        last_body = ""
        response_action = None

        for attempt in range(1, max_retries + 1):
            response_action = requests.post(URL, json=body_action, headers=headers)
            try:
                response_action.raise_for_status()
                last_error = ""
                last_status_code = ""
                last_body = ""
                break
            except requests.HTTPError:
                last_status_code = str(response_action.status_code)
                last_error = f"HTTP {response_action.status_code}"
                try:
                    last_body = response_action.text
                except Exception:
                    last_body = ""
                if attempt < max_retries:
                    print(
                        "Erro na requisicao da action "
                        f"({last_error}). Tentando novamente ({attempt}/{max_retries})..."
                    )

        if last_error:
            return {
                "failed": {
                    "action_raw": action_raw,
                    "action_normalizada": action_normalizada,
                    "erro": last_error,
                    "status_code": last_status_code,
                    "response_body": last_body,
                }
            }

        data_action = response_action.json()
        resource_action = data_action.get("resource", {})
        items_action = resource_action.get("items", [])

        rows = []
        extras_keys = set()

        for item in items_action:
            contact = item.get("contact", {}).get("Identity")
            action_item = item.get("action")
            extras = item.get("extras") or {}
            extras_keys.update(extras.keys())

            info = action_info.get(action_raw, {})
            rows.append(
                {
                    "inc_status": ";".join(sorted(info.get("status", set()))),
                    "inc_status_code": ";".join(sorted(info.get("status_code", set()))),
                    "inc_erro_message": " | ".join(sorted(info.get("erro_message", set()))),
                    "inc_total_count": info.get("total_count", 0),
                    "inc_error_count": info.get("error_count", 0),
                    "storageDate": item.get("storageDate"),
                    "storageDateBR": storage_date_br(item.get("storageDate")),
                    "cpf": extras.get("cpf", ""),
                    "contactIdentity": contact or "",
                    "category": item.get("category"),
                    "action": action_item,
                    "extras": extras,
                }
            )

        return {
            "rows": rows,
            "extras_keys": extras_keys,
        }

    total_actions = len(unique_actions)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(fetch_action, action_raw): (idx, action_raw)
            for idx, action_raw in enumerate(unique_actions, start=1)
        }

        for future in concurrent.futures.as_completed(future_map):
            idx, _action_raw = future_map[future]
            print(f"[{idx}/{total_actions}] Consultando action...")
            result = future.result()
            failed = result.get("failed")
            if failed:
                failed_actions.append(failed)
                continue
            eventos_rows.extend(result.get("rows", []))
            extras_keys_eventos.update(result.get("extras_keys", set()))

    total_inconsistencias_api = sum(
        int(enriched["item"].get("count", 0) or 0) for enriched in items_enriquecidos
    )
    total_erros = sum(
        int(enriched["item"].get("count", 0) or 0)
        for enriched in items_enriquecidos
        if enriched.get("is_error")
    )
    taxa_erro_geral = (
        (total_erros / total_inconsistencias_api) * 100 if total_inconsistencias_api else 0.0
    )

    total_registros_api = len(eventos_rows)
    total_actions_distintas = len({row.get("action", "") for row in eventos_rows if row.get("action")})
    total_contatos_distintos = len(
        {row.get("contactIdentity", "") for row in eventos_rows if row.get("contactIdentity")}
    )
    total_cpfs_distintos = len({row.get("cpf", "") for row in eventos_rows if row.get("cpf")})

    top_actions_counter = Counter(
        row.get("action", "") for row in eventos_rows if row.get("action")
    )

    erros_por_status = {}
    for enriched in items_enriquecidos:
        if not enriched.get("is_error"):
            continue
        item = enriched["item"]
        count = int(item.get("count", 0) or 0)
        key = (
            str(enriched.get("status") or "sem status"),
            str(enriched.get("status_code") or "sem codigo"),
            limpar_mensagem(str(enriched.get("erro_message") or "sem mensagem")),
        )
        erros_por_status[key] = erros_por_status.get(key, 0) + count

    erros_por_status_ordenado = sorted(
        erros_por_status.items(), key=lambda kv: kv[1], reverse=True
    )

    categorias = {}
    for enriched in items_enriquecidos:
        item = enriched["item"]
        category = str(item.get("category") or "sem categoria")
        count = int(item.get("count", 0) or 0)
        info = categorias.setdefault(category, {"total": 0, "erros": 0})
        info["total"] += count
        if enriched.get("is_error"):
            info["erros"] += count

    categorias_ordenadas = sorted(
        categorias.items(), key=lambda kv: kv[1]["total"], reverse=True
    )

    eventos_por_dia = Counter()
    for row in eventos_rows:
        chave_data = date_key(row.get("storageDateBR") or row.get("storageDate") or "")
        if chave_data:
            eventos_por_dia[chave_data] += 1

    incons_por_dia = Counter()
    erros_por_dia = Counter()
    for enriched in items_enriquecidos:
        item = enriched["item"]
        chave_data = date_key(item.get("storageDate") or "")
        if not chave_data:
            continue
        count = int(item.get("count", 0) or 0)
        incons_por_dia[chave_data] += count
        if enriched.get("is_error"):
            erros_por_dia[chave_data] += count

    todas_datas = sorted(set(eventos_por_dia) | set(incons_por_dia) | set(erros_por_dia))

    wb = Workbook()

    ws_det = wb.active
    ws_det.title = "relatorio por eventos"

    ws_det.append(["Router", router_nome])
    ws_det.append(["Tracking", tracking])
    ws_det.append(["Data inicial", start_date])
    ws_det.append(["Data final", end_date])
    ws_det.append([])
    ws_det.append(
        [
            "storageDate",
            "category",
            "status",
            "statusCode",
            "erro_message",
            "count",
            "action_raw",
        ]
    )

    for enriched in items_enriquecidos:
        item = enriched["item"]
        ws_det.append(
            [
                item.get("storageDate"),
                item.get("category"),
                enriched.get("status"),
                enriched.get("status_code"),
                enriched.get("erro_message"),
                item.get("count", 0),
                item.get("action", ""),
            ]
        )

    ws_eventos = wb.create_sheet("relatorio por action")
    extras_keys_sorted = sorted(extras_keys_eventos)
    extras_headers = [f"Extras_{key}" for key in extras_keys_sorted]
    ws_eventos.append(
        [
            "inc_status",
            "inc_statusCode",
            "inc_erro_message",
            "inc_total_count",
            "inc_error_count",
            "storageDate",
            "storageDateBR",
            "cpf",
            "contactIdentity",
            "category",
            "action",
            *extras_headers,
        ]
    )

    for row in eventos_rows:
        extras = row.get("extras") or {}
        extras_values = []
        for key in extras_keys_sorted:
            value = extras.get(key, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            if value is None:
                value = ""
            extras_values.append(value)

        ws_eventos.append(
            [
                row.get("inc_status", ""),
                row.get("inc_status_code", ""),
                row.get("inc_erro_message", ""),
                row.get("inc_total_count", 0),
                row.get("inc_error_count", 0),
                row.get("storageDate", ""),
                row.get("storageDateBR", ""),
                row.get("cpf", ""),
                row.get("contactIdentity", ""),
                row.get("category", ""),
                row.get("action", ""),
                *extras_values,
            ]
        )

    ws_resumo = wb.create_sheet("resumo_geral")
    ws_resumo.append(["metrica", "valor"])
    ws_resumo.append(["router", router_nome])
    ws_resumo.append(["tracking", tracking])
    ws_resumo.append(["data_inicial", start_date])
    ws_resumo.append(["data_final", end_date])
    ws_resumo.append(["total_inconsistencias_api", total_inconsistencias_api])
    ws_resumo.append(["total_registros_api", total_registros_api])
    ws_resumo.append(["total_actions_distintas", total_actions_distintas])
    ws_resumo.append(["total_contatos_distintos", total_contatos_distintos])
    ws_resumo.append(["total_cpfs_distintos", total_cpfs_distintos])
    ws_resumo.append(["total_erros", total_erros])
    ws_resumo.append(["taxa_erro_geral_percent", round(taxa_erro_geral, 2)])
    ws_resumo.append(["actions_consulta_falharam", len(failed_actions)])

    ws_top = wb.create_sheet("top_actions")
    ws_top.append(
        [
            "action",
            "ocorrencias_api",
            "inc_total_count",
            "inc_error_count",
            "taxa_erro_percent",
        ]
    )

    for action, ocorrencias in top_actions_counter.most_common():
        info = action_info.get(action, {})
        inc_total = int(info.get("total_count", 0) or 0)
        inc_error = int(info.get("error_count", 0) or 0)
        taxa = (inc_error / inc_total) * 100 if inc_total else 0.0
        ws_top.append([action, ocorrencias, inc_total, inc_error, round(taxa, 2)])

    ws_err = wb.create_sheet("erros_por_status")
    ws_err.append(
        [
            "status",
            "statusCode",
            "erro_message",
            "ocorrencias",
            "percentual_total_erros",
        ]
    )

    for (status, status_code, mensagem), ocorrencias in erros_por_status_ordenado:
        percentual = (ocorrencias / total_erros) * 100 if total_erros else 0.0
        ws_err.append([status, status_code, mensagem, ocorrencias, round(percentual, 2)])

    ws_cat = wb.create_sheet("categorias")
    ws_cat.append(["category", "total_eventos", "total_erros", "taxa_erro_percent"])

    for category, info in categorias_ordenadas:
        total = info["total"]
        erros = info["erros"]
        taxa = (erros / total) * 100 if total else 0.0
        ws_cat.append([category, total, erros, round(taxa, 2)])

    ws_ts = wb.create_sheet("serie_temporal")
    ws_ts.append(
        [
            "data",
            "eventos_action_api",
            "inconsistencias_api",
            "erros_api",
            "taxa_erro_percent",
        ]
    )

    for dia in todas_datas:
        eventos = eventos_por_dia.get(dia, 0)
        inconsistencias = incons_por_dia.get(dia, 0)
        erros = erros_por_dia.get(dia, 0)
        taxa = (erros / inconsistencias) * 100 if inconsistencias else 0.0
        ws_ts.append([dia, eventos, inconsistencias, erros, round(taxa, 2)])

    ws_resumo_inc = wb.create_sheet("resumo_incons")
    ws_resumo_inc.append(["ocorrencias_totais", "erro_message"])
    for mensagem, total in erros_ordenados:
        ws_resumo_inc.append([total, limpar_mensagem(mensagem)])

    if failed_actions:
        ws_failed = wb.create_sheet("actions_falharam")
        ws_failed.append(
            ["action_raw", "action_normalizada", "erro", "status_code", "response_body"]
        )
        for item in failed_actions:
            ws_failed.append(
                [
                    item.get("action_raw", ""),
                    item.get("action_normalizada", ""),
                    item.get("erro", ""),
                    item.get("status_code", ""),
                    item.get("response_body", ""),
                ]
            )

    today_str = datetime.now().strftime("%Y%m%d")

    home_dir = Path.home()
    documentos_dir = home_dir / "Documentos"
    if not documentos_dir.exists():
        documentos_dir = home_dir / "Documents"

    relatorio_root = documentos_dir / "relatorio de eventos"
    tipo_subpasta = relatorio_root / "eventDetails" / router_dir
    today_folder = tipo_subpasta / f"relatorio_{today_str}"
    today_folder.mkdir(parents=True, exist_ok=True)

    base_name = f"relatorio_analitico_eventos_{router_dir}_{today_str}"
    ext = ".xlsx"
    file_name = base_name + ext
    file_path = today_folder / file_name

    version = 1
    while file_path.exists():
        version += 1
        file_name = f"{base_name}_v{version}{ext}"
        file_path = today_folder / file_name

    wb.save(file_path)

    print("\n===== ARQUIVO GERADO =====")
    print("Arquivo salvo como:", file_path)
    if failed_actions:
        print("\n===== RESUMO DE FALHAS =====")
        print(f"Actions com erro: {len(failed_actions)}")
        print("Confira a aba 'actions_falharam' no Excel para detalhes.")
    print("\nEncerrando o programa.\n")
    break
